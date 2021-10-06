# --- coding: utf-8 ---
# @Time    : 2021/10/6 9:33
# @Author  : XuJie
# @Email   : 472350338@qq.com
# @GitHub  : XuJieYa
# @File    : python_myexcel.py
# @Software: PyCharm
import pymysql
from openpyxl.reader.excel import load_workbook as lw

# 建立一个MySql链接
conn = pymysql.connect(
    host='localhost',
    user='root',
    password='root',
    db='test',
    port=3306,
    charset='utf8',
)
# 获得游标
cur = conn.cursor()
# 创建插入sql语句
query = 'insert into table_of_dishes(`菜品ID`, `菜品名称`, `食物编号`, `重量`, `冷or热菜`, `是or否主成分`) values (%s, %s, %s, %s, %s, %s)'

Workbook = lw("Data_sheet_one.xlsx")
Worksheet = Workbook[Workbook.sheetnames[0]]
rows = Worksheet.max_row
columns = Worksheet.max_column

try:
    for i in range(2, Worksheet.max_row):
        values = (Worksheet.cell(i, 1).value, Worksheet.cell(i, 2).value, Worksheet.cell(i, 1).value, Worksheet.cell(i, 6).value, Worksheet.cell(i, 3).value, Worksheet.cell(i, 5).value)
        # 执行语句
        cur.execute(query, values)
except Exception as error:
    print(error)
cur.close()
conn.commit()
conn.close()
